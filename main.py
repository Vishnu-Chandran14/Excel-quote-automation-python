import os

print(f"cuurent folder:", os.getcwd())

import pandas as pd
from datetime import datetime


from openpyxl.styles import Font, Alignment, Border, Side



#read the excel , sheet_name = 0 means that read first sheet alone:
df = pd.read_excel("Input/HWM_Quote_file.xlsx", sheet_name=0, engine="openpyxl")

# #by using os create the output folder
# os.makedirs("Output", exist_ok=True)



for order, data in df.groupby("Order No."):
    # print(f"Total line items: {data.shape[0]}")  
    #--------------Header information-----------------#
    quote_number = order
    first_row = data.iloc[0]

    account_number = str(first_row["Last five # AC"])
    account_name = first_row["CUSTOMER NAME:"]
    Contact_name = first_row["CONTACT NAME:"]
    Contact_email = first_row["CONTACT EMAIL:"]
    contact_phone = first_row["CONTACT PHONE:"]

    start_date = pd.to_datetime(first_row["Billing Period Start Date"])
    formated_start_date = start_date.strftime("%m/%d/%Y")

    end_date = pd.to_datetime(first_row["Billing Period End Date"])
    formated_end_date = end_date.strftime("%m/%d/%Y")

    #--------------DEVICE TABLE--------------------------#

    row_list = []

    for index, row in df.iterrows():

        qty_Ordered = int(row["Qty Ordered"])

        first_row = {
            "QTY" : qty_Ordered,
           "DEVICE TYPE": row['Product No.'],
           "DEVICE DESCRIPTION": row['Product Name'],
           "SERIAL NUMBER" : '',
           }

        row_list.append(first_row)
    
        for qty in range(qty_Ordered - 1):

            blank_row = {
                "QTY" : '',
                "DEVICE TYPE": '',
               "DEVICE DESCRIPTION": '',
               "SERIAL NUMBER" : '',
               }
            row_list.append(blank_row)

        #add separator row line for each product line:

        row_list.append({
            "QTY" : '',
            "DEVICE TYPE": '',
            "DEVICE DESCRIPTION": '',
            "SERIAL NUMBER" : '',
            })


    device_df = pd.DataFrame(row_list)
    

    #---------------FILE NAME -------------------#
    file_name = f"Output/{order} Renewal Quote Device Listing {account_number}.xlsx"

    #--------------WRITE EXCEL ---------------------#

    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:

        device_df.to_excel(writer, sheet_name="Quote", startrow=14, index=False)
        sheet = writer.sheets["Quote"] #sheets is a dictionary, not a function. sheet = writer.sheets["Sheet1"]
        
        #------A, B, C, D columns width-----#

        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 22
        sheet.column_dimensions["C"].width = 60
        sheet.column_dimensions["D"].width = 20
        
        sheet.merge_cells("A1:D1")
        
        sheet["A1"] = "RENEWAL QUOTE DEVICE / SERIAL NUMBER LISTING"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A1"].alignment = Alignment(horizontal="center")


        sheet["A3"] = "RENEWAL QUOTE #:"
        sheet["A3"].font = Font(bold=True, size=11)
        sheet["B3"] = int(quote_number)
        sheet["B3"].alignment = Alignment(horizontal="left")

        sheet["A5"] = "CUSTOMER #:"
        sheet["A5"].font = Font(bold=True, size=11)
        sheet["B5"] = account_number
        sheet["B5"].alignment = Alignment(horizontal="left")

        sheet["A6"] = "CUSTOMER NAME:"
        sheet["A6"].font = Font(bold=True, size=11)
        sheet["B6"] = account_name
        sheet["B6"].alignment = Alignment(horizontal="left")

        sheet["A7"] = "CONTACT NAME:"
        sheet["A7"].font = Font(bold=True, size=11)
        sheet["B7"] = Contact_name
        sheet["B7"].alignment = Alignment(horizontal="left")

        sheet["A8"] = "CONTACT EMAIL:"
        sheet["A8"].font = Font(bold=True, size=11)
        sheet["B8"] = Contact_email
        sheet["B8"].alignment = Alignment(horizontal="left")

        sheet["A9"] = "CONTACT PHONE:"
        sheet["A9"].font = Font(bold=True, size=11)
        sheet["B9"] = int(contact_phone)
        sheet["B9"].alignment = Alignment(horizontal="left")

        sheet["A10"] = "START DATE:"
        sheet["A10"].font = Font(bold=True, size=11)
        sheet["B10"] = formated_start_date
        sheet["B10"].alignment = Alignment(horizontal="left")

        sheet["A11"] = "END DATE:"
        sheet["A11"].font = Font(bold=True, size=11)
        sheet["B11"] = formated_end_date
        sheet["B11"].alignment = Alignment(horizontal="left")



        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'))
        
        for row in sheet.iter_rows(min_row=15, max_row=sheet.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = thin_border



    print(f"Quote #{quote_number} sheet has been created successfully!")





